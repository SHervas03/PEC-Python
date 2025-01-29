import openpyxl
import pandas as pd
import xlwt

def aaa(df_default):
    df_default['Descripción'] = df_default['Descripción'].fillna('No description')
    df_default['Servicio'] = df_default['Servicio'].fillna('Unspecified')
    df_default['Torre'] = df_default['Torre'].fillna('Unspecified')
    df_default['Entorno'] = df_default['Entorno'].fillna('Unspecified')
    df_default['Estado cumplimiento SLA'] = df_default['Estado cumplimiento SLA'].fillna('Unspecified')
    return df_default

def bbb(df, df_new_path):
    df.to_excel(df_new_path, index=False)


def ccc(df):
    print(df.info())

def ddd(new_file_path, xls_file):
    wb_xlsx = openpyxl.load_workbook(new_file_path)
    sheet_xlsx = wb_xlsx.active
    wb_xls = xlwt.Workbook()
    sheet_xls = wb_xls.add_sheet('Hoja 1')
    for row_idx, row in enumerate(sheet_xlsx.iter_rows(), start=0):
        for col_idx, cell in enumerate(row, start=0):
            sheet_xls.write(row_idx, col_idx, cell.value)

    wb_xls.save(xls_file)
    print("Sva")


if __name__ == "__main__":
    df_default = "./Data/Dataset.xls"
    df_new =  './Data/New_Dataset.xlsx'
    xls_file = './Data/New_Dataset.xls'
    df_dflt = pd.read_excel(df_default)

    # ccc(df_dflt)
    # df = aaa(df_dflt)
    # ccc(df)
    # bbb(df, df_new)
    # df_nw = pd.read_excel(df_new)
    # ccc(df_nw)
    ddd(df_new, xls_file)
    df = pd.read_excel(xls_file)
    print(df.info())



